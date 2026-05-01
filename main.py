import threading
from typing import Any, TypedDict, cast

import requests
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

URL_CONSULTA = "https://portal-sitram.sefaz.ce.gov.br/api-nota/notafiscal/por-chave-de-acesso/"
URL_LANCAMENTOS = "https://portal-sitram.sefaz.ce.gov.br/api-nota/notafiscal/lancamentos-nota-fiscal/"
URL_ITENS = "https://portal-sitram.sefaz.ce.gov.br/api-nota/notafiscal/itens-nota-fiscal/"

ARQUIVO_PLANILHA = "chaves.xlsx"
ARQUIVO_TXT = "chaves.txt"
ARQUIVO_SAIDA = "Relatorio_SITRAM.xlsx"
TIMEOUT = 10
MAX_WORKERS = 10

RENOMEAR_COLUNAS_NOTAS = {
    "chave": "Chave de Acesso",
    "status_consulta": "Status da Consulta",
    "id_nota": "ID da Nota",
    "numero_nf": "Número NF",
    "data_emissao": "Data de Emissão",
    "data_inclusao": "Data de Inclusão",
    "data_fato_gerador": "Data Fato Gerador",
    "selada": "Selada",
    "emitente": "Emitente",
    "codigo_emitente": "Código Emitente",
    "uf_emitente": "UF Emitente",
    "codigo_destinatario": "Código Destinatário",
    "valor_nota": "Valor Nota",
    "total_produtos": "Total Produtos",
    "base_calculo": "Base de Cálculo",
    "valor_icms": "Valor ICMS",
    "valor_ipi": "Valor IPI",
    "situacao": "Situação",
    "situacao_alteracao": "Situação Alteração",
    "situacao_do_imposto": "Situação do Imposto",
    "acao_fiscal_situacao_descricao": "Ação Fiscal",
    "id_acao_fiscal": "ID Ação Fiscal",
    "descricao_orgao_local": "Órgão Local",
    "pode_ser_alterada": "Pode Ser Alterada",
    "retorno": "Retorno",
    "sta_alteracao": "STA Alteração",
    "numero_serie": "Número Série",
}

RENOMEAR_COLUNAS_LANCAMENTOS = {
    "chave": "Chave de Acesso",
    "id_nota": "ID da Nota",
    "id_lancamento": "ID Lançamento",
    "id_lancamento_front": "ID Lançamento Front",
    "vencimento": "Vencimento",
    "valor": "Valor",
    "valor_pago": "Valor Pago",
    "situacao_codigo": "Código Situação",
    "situacao": "Situação",
    "tipo_credenciamento": "Tipo Credenciamento",
    "tipo_credenciamento_descricao": "Credenciamento Descrição",
    "codigo": "Código",
    "descricao_abreviada": "Descrição Abreviada",
    "descricao": "Descrição",
    "api_lancamentos": "API Lançamentos",
}

RENOMEAR_COLUNAS_ITENS = {
    "chave": "Chave de Acesso",
    "id_nota": "ID da Nota",
    "id_item": "ID Item",
    "item": "Item",
    "codigo_produto": "Código Produto",
    "descricao_produto": "Descrição Produto",
    "quantidade": "Quantidade",
    "valor_item": "Valor Item",
    "valor_icms_destacado": "Valor ICMS Destacado",
    "icms": "ICMS",
    "valor_ipi": "Valor IPI",
    "valor_fecop": "Valor FECOP",
    "valor_unitario": "Valor Unitário",
    "valor_aliquota": "Valor Alíquota",
    "valor_bc": "Valor Base de Cálculo",
    "indicador_insumo": "Indicador Insumo",
    "indicador_consumo": "Indicador Consumo",
    "indicador_ativo_fixo": "Indicador Ativo Fixo",
    "codigo_csta": "Código CSTA",
    "codigo_cstb": "Código CSTB",
    "cfop": "CFOP",
    "cfop_descricao": "Descrição CFOP",
    "ncm": "NCM",
    "ncm_descricao": "Descrição NCM",
    "api_itens": "API Itens",
}


class NotaRegistro(TypedDict):
    chave: str
    status_consulta: str
    api_consulta: str
    id_nota: str | None
    numero_nf: int | None
    data_emissao: str | None
    data_inclusao: str | None
    data_fato_gerador: str | None
    selo_nf: str | None
    selada: bool | None
    emitente: str | None
    codigo_emitente: str | None
    uf_emitente: str | None
    destinatario: str | None
    codigo_destinatario: int | None
    uf_destinatario: str | None
    tipo_emitente: int | None
    tipo_emitente_descricao: str | None
    tipo_destinatario: int | None
    tipo_destinatario_descricao: str | None
    valor_nota: float
    total_produtos: float
    base_calculo: float
    valor_icms: float
    valor_ipi: float
    situacao: str | None
    situacao_codigo: int | None
    situacao_transito_livre: int | None
    situacao_transito_livre_descricao: str | None
    situacao_alteracao: str | None
    tipo_alteracao_contribuinte: int | None
    situacao_do_imposto: str | None
    acao_fiscal_situacao_descricao: str | None
    id_acao_fiscal: str | None
    descricao_orgao_local: str | None
    nome_transportadora: str | None
    orgao_local_evento_sigla: str | None
    orgao_local_evento_descricao: str | None
    pode_ser_alterada: bool | None
    retorno: bool | None
    sta_alteracao: bool | None
    numero_serie: str | None
    data_consulta: str


class LancamentoRegistro(TypedDict):
    chave: str
    id_nota: str
    id_lancamento: int | None
    id_lancamento_front: str | None
    vencimento: str | None
    valor: float
    valor_pago: float
    situacao_codigo: int | None
    situacao: str | None
    tipo_credenciamento: int | None
    tipo_credenciamento_descricao: str | None
    codigo: str | None
    descricao_abreviada: str | None
    descricao: str | None
    api_lancamentos: str


class ItemRegistro(TypedDict):
    chave: str
    id_nota: str
    id_item: str | None
    item: int | str | None
    codigo_produto: str | None
    descricao_produto: str | None
    quantidade: float
    valor_item: float
    valor_icms_destacado: float
    icms: float
    valor_ipi: float
    valor_fecop: float
    valor_unitario: float
    valor_aliquota: float
    valor_bc: float
    indicador_insumo: bool | None
    indicador_consumo: bool | None
    indicador_ativo_fixo: bool | None
    codigo_csta: str | None
    codigo_cstb: str | None
    cfop: str | None
    cfop_descricao: str | None
    ncm: str | None
    ncm_descricao: str | None
    api_itens: str


print("Iniciando a coleta de dados da API...")


def criar_sessao() -> requests.Session:
    session = requests.Session()
    retries = Retry(total=3, backoff_factor=0.3, status_forcelist=[500, 502, 503, 504])
    adapter = HTTPAdapter(pool_connections=100, pool_maxsize=100, max_retries=retries)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


def localizar_arquivo_entrada() -> Path | None:
    planilha_padrao = Path(ARQUIVO_PLANILHA)
    if planilha_padrao.exists():
        return planilha_padrao

    for arquivo in sorted(Path(".").glob("*.xlsx")):
        if arquivo.name != ARQUIVO_SAIDA:
            return arquivo

    arquivo_txt = Path(ARQUIVO_TXT)
    if arquivo_txt.exists():
        return arquivo_txt

    return None


def carregar_chaves() -> list[str]:
    arquivo = localizar_arquivo_entrada()

    if arquivo is None:
        raise FileNotFoundError("Nenhum arquivo encontrado. Use chaves.xlsx ou chaves.txt.")

    if arquivo.suffix.lower() in [".xlsx", ".xls"]:
        print(f"Lendo planilha: {arquivo.name}")
        df = pd.read_excel(arquivo, dtype=str)
        coluna_chave = next((col for col in df.columns if "chave" in col.lower()), df.columns[0])
        valores = df[coluna_chave].dropna().astype(str).tolist()
    else:
        print(f"Lendo arquivo texto: {arquivo.name}")
        valores = arquivo.read_text(encoding="utf-8").splitlines()

    chaves = []
    for valor in valores:
        chave = "".join(filter(str.isdigit, str(valor).strip()))
        if chave:
            chaves.append(chave)

    if not chaves:
        raise ValueError("Nenhuma chave válida foi encontrada no arquivo de entrada.")

    return chaves


def normalizar_valor(valor: Any) -> float:
    try:
        if valor in (None, ""):
            return 0.0
        return float(valor)
    except (TypeError, ValueError):
        return 0.0


def formatar_data_brasileira(valor: Any) -> str | None:
    if valor in (None, ""):
        return None

    texto = str(valor).strip()
    formatos = [
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y",
    ]

    for formato in formatos:
        try:
            return datetime.strptime(texto, formato).strftime("%d/%m/%Y")
        except ValueError:
            continue

    return texto


def consultar_nota(chave: str) -> tuple[list[NotaRegistro], list[LancamentoRegistro], list[ItemRegistro]]:
    session = criar_sessao()
    chave = "".join(filter(str.isdigit, str(chave)))

    notas: list[NotaRegistro] = []
    lancamentos_lista: list[LancamentoRegistro] = []
    itens_lista: list[ItemRegistro] = []

    nota_base: NotaRegistro = {
        "chave": chave,
        "status_consulta": "CHAVE INVÁLIDA",
        "api_consulta": URL_CONSULTA,
        "id_nota": None,
        "numero_nf": None,
        "data_emissao": None,
        "data_inclusao": None,
        "data_fato_gerador": None,
        "selo_nf": None,
        "selada": None,
        "emitente": None,
        "codigo_emitente": None,
        "uf_emitente": None,
        "destinatario": None,
        "codigo_destinatario": None,
        "uf_destinatario": None,
        "tipo_emitente": None,
        "tipo_emitente_descricao": None,
        "tipo_destinatario": None,
        "tipo_destinatario_descricao": None,
        "valor_nota": 0.0,
        "total_produtos": 0.0,
        "base_calculo": 0.0,
        "valor_icms": 0.0,
        "valor_ipi": 0.0,
        "situacao": None,
        "situacao_codigo": None,
        "situacao_transito_livre": None,
        "situacao_transito_livre_descricao": None,
        "situacao_alteracao": None,
        "tipo_alteracao_contribuinte": None,
        "situacao_do_imposto": None,
        "acao_fiscal_situacao_descricao": None,
        "id_acao_fiscal": None,
        "descricao_orgao_local": None,
        "nome_transportadora": None,
        "orgao_local_evento_sigla": None,
        "orgao_local_evento_descricao": None,
        "pode_ser_alterada": None,
        "retorno": None,
        "sta_alteracao": None,
        "numero_serie": None,
        "data_consulta": datetime.now().strftime("%d/%m/%Y %H:%M"),
    }

    if len(chave) != 44:
        notas.append(nota_base)
        return notas, lancamentos_lista, itens_lista

    try:
        r = session.get(URL_CONSULTA + chave, timeout=TIMEOUT)

        if r.status_code != 200:
            nota_base["status_consulta"] = f"ERRO CONSULTA API {r.status_code}"
            notas.append(nota_base)
            return notas, lancamentos_lista, itens_lista

        resposta_json = r.json()
        conteudo = resposta_json.get("content", [])

        if not conteudo:
            nota_base["status_consulta"] = "NF SEM RETORNO NO SITRAM"
            notas.append(nota_base)
            return notas, lancamentos_lista, itens_lista

        nota = conteudo[0]
        id_nota = str(nota.get("id")) if nota.get("id") is not None else None

        nota_base.update(
            {
                "status_consulta": "SUCESSO",
                "id_nota": id_nota,
                "numero_nf": nota.get("numero"),
                "data_emissao": formatar_data_brasileira(nota.get("dataEmissao")),
                "data_inclusao": formatar_data_brasileira(nota.get("dataInclusao")),
                "data_fato_gerador": formatar_data_brasileira(nota.get("dataFatoGerador")),
                "selo_nf": nota.get("numeroSelo"),
                "selada": nota.get("selada"),
                "emitente": nota.get("nomeEmitente"),
                "codigo_emitente": str(nota.get("codigoEmitente")).zfill(14) if nota.get("codigoEmitente") is not None else None,
                "uf_emitente": nota.get("ufEmitente"),
                "destinatario": nota.get("nomeDestinatario"),
                "codigo_destinatario": nota.get("codigoDestinatario"),
                "uf_destinatario": nota.get("ufDestinatario"),
                "tipo_emitente": nota.get("tipoEmitente"),
                "tipo_emitente_descricao": nota.get("tipoEmitenteDescricao"),
                "tipo_destinatario": nota.get("tipoDestinatario"),
                "tipo_destinatario_descricao": nota.get("tipoDestinatarioDescricao"),
                "valor_nota": normalizar_valor(nota.get("valorTotalNota") or nota.get("total")),
                "total_produtos": normalizar_valor(nota.get("totalProdutos")),
                "base_calculo": normalizar_valor(nota.get("baseCalculo")),
                "valor_icms": normalizar_valor(nota.get("valorIcmsTotal")),
                "valor_ipi": normalizar_valor(nota.get("valorIpi")),
                "situacao": nota.get("situacaoDescricao"),
                "situacao_codigo": nota.get("situacao"),
                "situacao_transito_livre": nota.get("situacaoTransitoLivre"),
                "situacao_transito_livre_descricao": nota.get("situacaoTransitoLivreDescricao"),
                "situacao_alteracao": nota.get("situacaoAlteracao"),
                "tipo_alteracao_contribuinte": nota.get("tipoAlteracaoContribuinte"),
                "situacao_do_imposto": nota.get("situacaoDoImposto"),
                "acao_fiscal_situacao_descricao": nota.get("acaoFiscalSituacaoDescricao"),
                "id_acao_fiscal": nota.get("idAcaoFiscal"),
                "descricao_orgao_local": nota.get("descricaoOrgaoLocal"),
                "nome_transportadora": nota.get("nomeTransportadora"),
                "orgao_local_evento_sigla": nota.get("orgaoLocalEventoSigla"),
                "orgao_local_evento_descricao": nota.get("orgaoLocalEventoDescricao"),
                "pode_ser_alterada": nota.get("podeSerAlterada"),
                "retorno": nota.get("retorno"),
                "sta_alteracao": nota.get("staAlteracao"),
                "numero_serie": nota.get("numeroSerie"),
            }
        )
        notas.append(cast(NotaRegistro, nota_base.copy()))

        if id_nota is not None:
            rl = session.get(URL_LANCAMENTOS + str(id_nota), timeout=TIMEOUT)
            if rl.status_code != 200:
                print(f"[AVISO] Nota {id_nota} - lançamentos: status HTTP {rl.status_code} (chave: {chave})")
            if rl.status_code == 200:
                lancamentos_json = rl.json()
                if isinstance(lancamentos_json, dict):
                    lancamentos_json = lancamentos_json.get("content", [])
                if not isinstance(lancamentos_json, list) or len(lancamentos_json) == 0:
                    print(f"[AVISO] Nota {id_nota} - lançamentos: sem dados retornados (chave: {chave})")
                if isinstance(lancamentos_json, list):
                    for lanc in lancamentos_json:
                        lancamentos_lista.append(
                            {
                                "chave": chave,
                                "id_nota": id_nota,
                                "id_lancamento": lanc.get("id"),
                                "id_lancamento_front": lanc.get("idLancamentoFront"),
                                "vencimento": formatar_data_brasileira(lanc.get("vencimento")),
                                "valor": normalizar_valor(lanc.get("valor")),
                                "valor_pago": normalizar_valor(lanc.get("valorPago")),
                                "situacao_codigo": lanc.get("situacao"),
                                "situacao": lanc.get("siuacaoDescricao") or lanc.get("situacaoDescricao"),
                                "tipo_credenciamento": lanc.get("tipoCredenciamento"),
                                "tipo_credenciamento_descricao": lanc.get("tipoCredenciamentoDescricao"),
                                "codigo": lanc.get("codigo"),
                                "descricao_abreviada": lanc.get("descricaoAbreviada"),
                                "descricao": lanc.get("descricao"),
                                "api_lancamentos": URL_LANCAMENTOS,
                            }
                        )

            ri = session.get(URL_ITENS + str(id_nota), timeout=TIMEOUT)
            if ri.status_code == 200:
                itens_json = ri.json()
                conteudo_itens = itens_json.get("content", []) if isinstance(itens_json, dict) else itens_json
                if isinstance(conteudo_itens, list):
                    for item in conteudo_itens:
                        itens_lista.append(
                            {
                                "chave": chave,
                                "id_nota": id_nota,
                                "id_item": item.get("id"),
                                "item": item.get("numero") or item.get("numeroItem") or item.get("item"),
                                "codigo_produto": item.get("codigoProduto") or item.get("codigo"),
                                "descricao_produto": item.get("descricaoProduto") or item.get("descricao"),
                                "quantidade": normalizar_valor(item.get("quantidade") or item.get("quantidadeComercial")),
                                "valor_item": normalizar_valor(item.get("valorItem") or item.get("valorTotal") or item.get("valorTotalBrutoProduto")),
                                "valor_icms_destacado": normalizar_valor(item.get("valorIcmsDestacado")),
                                "icms": normalizar_valor(item.get("icms")),
                                "valor_ipi": normalizar_valor(item.get("valorIPI")),
                                "valor_fecop": normalizar_valor(item.get("valorFecop")),
                                "valor_unitario": normalizar_valor(item.get("valorUnitario")),
                                "valor_aliquota": normalizar_valor(item.get("valorAliquota")),
                                "valor_bc": normalizar_valor(item.get("valorBc")),
                                "indicador_insumo": item.get("indicadorInsumo"),
                                "indicador_consumo": item.get("indicadorConsumo"),
                                "indicador_ativo_fixo": item.get("indicadorAtivoFixo"),
                                "codigo_csta": item.get("codigoCSTA"),
                                "codigo_cstb": item.get("codigoCSTB"),
                                "cfop": item.get("cfop"),
                                "cfop_descricao": item.get("cfopDescricao"),
                                "ncm": item.get("ncm"),
                                "ncm_descricao": item.get("ncmDescricao"),
                                "api_itens": URL_ITENS,
                            }
                        )


    except requests.Timeout:
        nota_base["status_consulta"] = "TIMEOUT NA CONSULTA"
        notas = [nota_base]
    except requests.RequestException as erro:
        nota_base["status_consulta"] = f"FALHA DE CONEXÃO: {erro}"
        notas = [nota_base]
    except ValueError:
        nota_base["status_consulta"] = "RESPOSTA JSON INVÁLIDA"
        notas = [nota_base]

    return notas, lancamentos_lista, itens_lista


def consultar_notas_em_paralelo(
    chaves: list[str],
    progress_callback: Any = None,
    stop_event: threading.Event | None = None,
) -> tuple[list[NotaRegistro], list[LancamentoRegistro], list[ItemRegistro]]:
    notas_total: list[NotaRegistro] = []
    lancamentos_total: list[LancamentoRegistro] = []
    itens_total: list[ItemRegistro] = []
    total = len(chaves)
    concluidas = 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        tarefas = [executor.submit(consultar_nota, chave) for chave in chaves]

        for future in as_completed(tarefas):
            if stop_event is not None and stop_event.is_set():
                for t in tarefas:
                    t.cancel()
                print("[AVISO] Processamento cancelado pelo usuário.")
                break
            notas, lancamentos, itens = future.result()
            notas_total.extend(notas)
            lancamentos_total.extend(lancamentos)
            itens_total.extend(itens)
            concluidas += 1
            if progress_callback is not None:
                progress_callback(concluidas, total)

    return notas_total, lancamentos_total, itens_total


def aplicar_estilo(ws: Any) -> None:
    azul = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    branco = Font(color="FFFFFF", bold=True)
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    formato_contabil = '_-R$ * #,##0.00_-;[Red]-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'

    cabecalhos = [cell.value for cell in ws[1]]
    ws.freeze_panes = "A2"

    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = azul
        cell.font = branco
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, coluna in enumerate(ws.columns, start=1):
        maior = 0
        letra = get_column_letter(col_idx)
        nome_coluna = str(cabecalhos[col_idx - 1]).lower() if col_idx - 1 < len(cabecalhos) else ""
        for cell in coluna:
            cell.border = borda
            if cell.row > 1:
                if isinstance(cell.value, (int, float)) and any(chave in nome_coluna for chave in ["valor", "icms", "ipi", "fecop", "total", "base"]):
                    cell.number_format = formato_contabil
                elif any(chave in nome_coluna for chave in ["data", "emissão", "inclusão", "vencimento", "fato"]):
                    cell.alignment = Alignment(horizontal="center")
            if cell.value not in (None, ""):
                maior = max(maior, len(str(cell.value)))
        ws.column_dimensions[letra].width = min(maior + 3, 60)


def gerar_excel_profissional(
    notas: list[NotaRegistro],
    lancamentos: list[LancamentoRegistro],
    itens: list[ItemRegistro],
    nome_arquivo: str = ARQUIVO_SAIDA,
) -> None:
    df_notas = pd.DataFrame(notas).rename(columns=RENOMEAR_COLUNAS_NOTAS)
    colunas_notas = [c for c in RENOMEAR_COLUNAS_NOTAS.values() if c in df_notas.columns]
    df_notas = df_notas[colunas_notas]
    df_lancamentos = pd.DataFrame(lancamentos).rename(columns=RENOMEAR_COLUNAS_LANCAMENTOS)
    df_itens = pd.DataFrame(itens).rename(columns=RENOMEAR_COLUNAS_ITENS)
    df_apis = pd.DataFrame(
        [
            {"api": "consulta", "url": URL_CONSULTA},
            {"api": "lancamentos", "url": URL_LANCAMENTOS},
            {"api": "itens", "url": URL_ITENS},
        ]
    )

    with pd.ExcelWriter(nome_arquivo, engine="openpyxl") as writer:
        df_notas.to_excel(writer, index=False, sheet_name="Notas")
        (df_lancamentos if not df_lancamentos.empty else pd.DataFrame([{"mensagem": "Sem lançamentos retornados"}])).to_excel(writer, index=False, sheet_name="Lancamentos")
        (df_itens if not df_itens.empty else pd.DataFrame([{"mensagem": "Sem itens retornados"}])).to_excel(writer, index=False, sheet_name="Itens")
        df_apis.to_excel(writer, index=False, sheet_name="APIs")

        wb = writer.book
        aplicar_estilo(writer.sheets["Notas"])
        aplicar_estilo(writer.sheets["Lancamentos"])
        aplicar_estilo(writer.sheets["Itens"])
        aplicar_estilo(writer.sheets["APIs"])

        resumo = wb.create_sheet("Resumo")
        resumo.freeze_panes = "A2"
        resumo["A1"] = "KPIs de Processamento"
        resumo["A1"].font = Font(size=14, bold=True)

        total_notas = len(df_notas)
        total_sucesso = int((df_notas["Status da Consulta"] == "SUCESSO").sum()) if not df_notas.empty and "Status da Consulta" in df_notas else 0
        valor_total = pd.to_numeric(df_notas.get("Valor Nota", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()
        total_lancamentos = len(df_lancamentos)
        total_itens = len(df_itens)

        metricas = [
            ["Total de Notas", total_notas],
            ["Notas com Sucesso", total_sucesso],
            ["Total de Lançamentos", total_lancamentos],
            ["Total de Itens", total_itens],
            ["Valor Total Notas", float(valor_total)],
            ["Data do Relatório", datetime.now().strftime("%d/%m/%Y %H:%M")],
        ]

        for linha, (titulo, valor) in enumerate(metricas, start=3):
            resumo.cell(row=linha, column=1, value=titulo).font = Font(bold=True)
            resumo.cell(row=linha, column=2, value=valor)
            if "Valor" in titulo:
                resumo.cell(row=linha, column=2).number_format = '"R$" #,##0.00'

        resumo.column_dimensions["A"].width = 24
        resumo.column_dimensions["B"].width = 20

    print(f"Relatório gerado com sucesso: {nome_arquivo}")


def main() -> None:
    chaves = carregar_chaves()
    print(f"Consultando {len(chaves)} chave(s)...")
    notas, lancamentos, itens = consultar_notas_em_paralelo(chaves)
    gerar_excel_profissional(notas, lancamentos, itens)


if __name__ == "__main__":
    main()
